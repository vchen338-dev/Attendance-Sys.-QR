import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
from pyzbar.pyzbar import decode
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import os
import cv2


names = [
    "Kin Acoba", "Christian Aresta", "Jayson Bambilla", "Ace Battung", "Marwin Caderao",
    "Marvin Contillo", "Simon Durwin", "Deo Egipto", "Rowel Estanislao", "Jochen Galera",
    "Ralph Ganacia", "Jareld Narag", "Rodel Peratir", "Norwin Ramos", "James Sante",
    "Edrian Suriaga", "Rolly Torida", "Xyza Bago", "Kate Pagaduan", "Aila Maggay",
    "Mary Rose PIllos", "Jessica San Juan", "Frances Talosig"
]


root = tk.Tk()
root.title("TVL ICT 12 QR Attendance")
root.attributes("-fullscreen", True)


tk.Label(root, text="TVL ICT 12", font=("Arial Black", 40, "bold")).pack(pady=10)
tk.Label(root, text="QR Attendance", font=("Arial", 20)).pack(pady=5)


video_label = tk.Label(root)
video_label.pack()


mode_var = tk.StringVar(value="AM Time In")
ttk.Label(root, text="Select Mode:", font=("Arial", 18)).pack()
ttk.Combobox(root, textvariable=mode_var, values=["AM Time In", "AM Time Out", "PM Time In", "PM Time Out"], font=("Arial", 14), state="readonly").pack(pady=10)


already_scanned = set()


def record_time(student_name, timestamp, event):
    file_path = os.path.join(os.path.expanduser('~'), 'Documents', 'attendance.xlsx')
    try:
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "AM Time In", "AM Time Out", "PM Time In", "PM Time Out", "Status"])
            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        row_found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == student_name:
                row_found = True
                if event == "AM Time In":
                    row[1].value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                elif event == "AM Time Out":
                    row[2].value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                elif event == "PM Time In":
                    row[3].value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                elif event == "PM Time Out":
                    row[4].value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                row[5].value = "Present"
                break

        if not row_found:
            row = ["", "", "", "", "", "Present"]
            row[0] = student_name
            idx = ["AM Time In", "AM Time Out", "PM Time In", "PM Time Out"].index(event)
            row[idx + 1] = timestamp.strftime('%Y-%m-%d %H:%M:%S')
            ws.append(row)

        wb.save(file_path)
        messagebox.showinfo("Success", f"{student_name} marked as {event}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to record attendance: {e}")


cap = cv2.VideoCapture(1)

def update_frame():
    ret, frame = cap.read()
    if not ret:
        return

    for code in decode(frame):
        qr_data = code.data.decode('utf-8')
        if qr_data in names and qr_data not in already_scanned:
            already_scanned.add(qr_data)
            record_time(qr_data, datetime.now(), mode_var.get())
            root.after(3000, lambda: already_scanned.discard(qr_data))  # Allow re-scan after delay

        (x, y, w, h) = code.rect
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 3)
        cv2.putText(frame, qr_data, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(img))
    video_label.imgtk = img
    video_label.configure(image=img)
    video_label.after(10, update_frame)

update_frame()


def export_to_excel(ws):
    try:
        wb = Workbook()
        new_ws = wb.active
        new_ws.title = "Attendance History"

        headers = ["Name", "AM Time In", "AM Time Out", "PM Time In", "PM Time Out", "Status"]
        fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)

        for col, head in enumerate(headers, 1):
            cell = new_ws.cell(row=1, column=col, value=head)
            cell.fill = fill
            cell.font = font

        row_index = 2
        for name in names:
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == name:
                    new_row = list(row)
                    status = ""
                    if row[1] and row[2]:
                        status += "AM Present"
                    elif row[1] and not row[2]:
                        status += "AM Cutting Class"
                    else:
                        status += "AM Absent"

                    if row[3] and row[4]:
                        status += ", PM Present"
                    elif row[3] and not row[4]:
                        status += ", PM Cutting Class"
                    else:
                        status += ", PM Absent"

                    new_row[5] = status
                    new_ws.append(new_row)
                    found = True
                    break

            if not found:
                new_ws.append([name, "", "", "", "", "AM Absent, PM Absent"])

            status_cell = new_ws.cell(row=row_index, column=6)
            if "Present" in status_cell.value and "Absent" not in status_cell.value:
                color = "00FF00"
            elif "Cutting Class" in status_cell.value:
                color = "FF0000"
            elif "Present" in status_cell.value:
                color = "FFA500"
            else:
                color = "FFFF00"

            for cell in new_ws[f"A{row_index}:F{row_index}"][0]:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row_index += 1

        file_name = datetime.now().strftime("attendance_history_%Y-%m-%d_%H-%M-%S.xlsx")
        file_path = os.path.join(os.path.expanduser('~'), 'Documents', file_name)
        wb.save(file_path)
        messagebox.showinfo("Exported", f"File saved at:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Export Error", str(e))


def reset_history():
    file_path = os.path.join(os.path.expanduser('~'), 'Documents', 'attendance.xlsx')
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            messagebox.showinfo("Reset", "Attendance history has been reset.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to reset: {e}")


def show_history():
    password = simpledialog.askstring("Admin Login", "Enter password:", show='*')
    if password != 'tvl2024':
        messagebox.showerror("Access Denied", "Incorrect password.")
        return

    file_path = os.path.join(os.path.expanduser('~'), 'Documents', 'attendance.xlsx')
    if not os.path.exists(file_path):
        messagebox.showinfo("No Data", "No attendance history found.")
        return

    history_window = tk.Toplevel(root)
    history_window.title("Attendance History")
    tk.Label(history_window, text="Attendance History", font=("Arial", 20, "bold")).pack(pady=10)

    frame = ttk.Frame(history_window)
    frame.pack(fill="both", expand=True)

    tree = ttk.Treeview(frame, columns=("Name", "AM Time In", "AM Time Out", "PM Time In", "PM Time Out", "Status"), show="headings")
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=200 if col != "Status" else 150)

    tree.pack(fill="both", expand=True)

    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

    ttk.Button(history_window, text="EXPORT TO EXCEL", command=lambda: export_to_excel(ws)).pack(pady=5)
    ttk.Button(history_window, text="RESET HISTORY", command=reset_history).pack(pady=5)


ttk.Button(root, text="ADMIN", command=show_history).pack(side="left", padx=20, pady=10)


style = ttk.Style()
style.configure("TButton", font=("Arial", 16), padding=10)


root.mainloop()
cap.release()
cv2.destroyAllWindows()
